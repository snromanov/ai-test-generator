"""
Экспортеры тест-кейсов в различные форматы с защитой от инъекций.
"""
import csv
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.agents.test_agent import GenerationResult, TestCase
from src.utils.logger import setup_logger

logger = setup_logger(__name__)


def sanitize_for_excel(text: str) -> str:
    """
    Защита от Formula Injection в Excel.
    
    Excel интерпретирует ячейки начинающиеся с =, +, -, @ как формулы.
    Это может привести к выполнению вредоносного кода.
    
    Args:
        text: Текст для санитизации
        
    Returns:
        Безопасный текст с префиксом ' если необходимо
    """
    if not text or not isinstance(text, str):
        return text
    
    # Опасные символы в начале ячейки
    dangerous_starts = ['=', '+', '-', '@', '\t', '\r', '\n']
    
    # Проверяем первый символ
    if text and text[0] in dangerous_starts:
        # Добавляем одинарную кавычку - Excel будет трактовать как текст
        logger.debug(f"Sanitized Excel formula injection attempt: {text[:50]}")
        return "'" + text
    
    return text


class Exporter(Protocol):
    """Протокол для экспортеров."""

    def export(self, results: list[GenerationResult], output_path: str) -> str:
        """Экспортирует результаты в файл."""
        ...


class ExcelExporter:
    """Экспортер в формат Excel (.xlsx)."""

    # Стили для Excel
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    PRIORITY_COLORS = {
        "Critical": "FF0000",
        "High": "FFA500",
        "Medium": "FFFF00",
        "Low": "90EE90"
    }

    # Цвета для компонентов
    COMPONENT_COLORS = {
        "backend": "4472C4",   # Синий
        "frontend": "ED7D31",  # Оранжевый
        "fullstack": "70AD47", # Зеленый
    }

    # Названия листов по слоям
    LAYER_SHEET_NAMES = {
        "api": "Tests-API",
        "ui": "Tests-UI",
        "integration": "Tests-Integration",
        "e2e": "Tests-E2E",
    }

    def export(
        self,
        results: list[GenerationResult],
        output_path: str,
        include_analysis: bool = True,
        group_by_layer: bool = False
    ) -> str:
        """
        Экспортирует результаты в Excel файл.

        Args:
            results: Список результатов генерации
            output_path: Путь для сохранения
            include_analysis: Включить лист с анализом
            group_by_layer: Группировать тест-кейсы по слоям (api, ui, integration, e2e)

        Returns:
            Путь к созданному файлу
        """
        logger.info(f"Начало экспорта в Excel: {output_path}")
        wb = Workbook()

        if group_by_layer:
            # Создаем листы по слоям
            self._create_layered_sheets(wb, results)
        else:
            # Создаем единый лист с тест-кейсами
            ws_tests = wb.active
            ws_tests.title = "Test Cases"
            self._create_test_cases_sheet(ws_tests, results)

        # Создаем лист с анализом если требуется
        if include_analysis:
            ws_analysis = wb.create_sheet("Requirements Analysis")
            self._create_analysis_sheet(ws_analysis, results)

        # Создаем сводный лист
        ws_summary = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(ws_summary, results, group_by_layer)

        # Сохраняем файл
        path = Path(output_path)
        if not path.suffix:
            path = path.with_suffix(".xlsx")

        wb.save(path)
        logger.info(f"Excel файл сохранен: {path}")
        return str(path)

    def _create_layered_sheets(self, wb: Workbook, results: list[GenerationResult]):
        """Создает отдельные листы для каждого слоя тестирования."""
        # Группируем тест-кейсы по слоям
        tests_by_layer: dict[str, list[tuple[TestCase, str]]] = {
            "api": [],
            "ui": [],
            "integration": [],
            "e2e": [],
        }

        for result in results:
            req_text = result.requirement_text[:100] + "..." if len(result.requirement_text) > 100 else result.requirement_text
            for tc in result.test_cases:
                layer = getattr(tc, 'layer', 'api') or 'api'
                if layer in tests_by_layer:
                    tests_by_layer[layer].append((tc, req_text))
                else:
                    tests_by_layer['api'].append((tc, req_text))

        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Создаем листы для каждого слоя с тестами
        for layer, tests in tests_by_layer.items():
            if tests:
                sheet_name = self.LAYER_SHEET_NAMES.get(layer, f"Tests-{layer.upper()}")
                ws = wb.create_sheet(sheet_name)
                self._create_layer_sheet(ws, tests, layer)

    def _create_layer_sheet(self, ws, tests: list[tuple[TestCase, str]], layer: str):
        """Создает лист с тест-кейсами для конкретного слоя."""
        headers = [
            "ID", "Название", "Приоритет", "Компонент",
            "Предусловия", "Шаги", "Ожидаемый результат",
            "Тип теста", "Техника", "Теги", "Требование"
        ]

        # Заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Данные
        row = 2
        for tc, req_text in tests:
            self._write_extended_test_case_row(ws, row, tc, req_text)
            row += 1

        # Настраиваем ширину колонок
        column_widths = [10, 40, 12, 12, 30, 50, 40, 15, 25, 20, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Автофильтр
        ws.auto_filter.ref = ws.dimensions

    def _write_extended_test_case_row(self, ws, row: int, tc: TestCase, req_text: str):
        """Записывает строку тест-кейса с расширенными полями."""
        # Форматируем предусловия
        preconditions = "\n".join(f"- {p}" for p in tc.preconditions) if tc.preconditions else ""

        # Форматируем шаги
        steps = "\n".join(f"{s['step']}. {s['action']}" for s in tc.steps) if tc.steps else ""

        # Форматируем теги
        tags = ", ".join(getattr(tc, 'tags', []) or [])

        # Получаем компонент
        component = getattr(tc, 'component', 'fullstack') or 'fullstack'

        # Санитизируем все значения
        values = [
            sanitize_for_excel(tc.id),
            sanitize_for_excel(tc.title),
            sanitize_for_excel(tc.priority),
            sanitize_for_excel(component),
            sanitize_for_excel(preconditions),
            sanitize_for_excel(steps),
            sanitize_for_excel(tc.expected_result),
            sanitize_for_excel(tc.test_type),
            sanitize_for_excel(tc.technique),
            sanitize_for_excel(tags),
            sanitize_for_excel(req_text)
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = self.CELL_ALIGNMENT
            cell.border = self.THIN_BORDER

            # Подсветка приоритета
            if col == 3 and value in self.PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=self.PRIORITY_COLORS[value],
                    end_color=self.PRIORITY_COLORS[value],
                    fill_type="solid"
                )

            # Подсветка компонента
            if col == 4 and value in self.COMPONENT_COLORS:
                cell.fill = PatternFill(
                    start_color=self.COMPONENT_COLORS[value],
                    end_color=self.COMPONENT_COLORS[value],
                    fill_type="solid"
                )
                cell.font = Font(color="FFFFFF")

    def _create_test_cases_sheet(self, ws, results: list[GenerationResult]):
        """Создает лист с тест-кейсами."""
        headers = [
            "ID", "Название", "Приоритет", "Предусловия",
            "Шаги", "Ожидаемый результат", "Тип теста", "Техника", "Требование"
        ]

        # Заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Данные
        row = 2
        for result in results:
            req_text = result.requirement_text[:100] + "..." if len(result.requirement_text) > 100 else result.requirement_text

            for tc in result.test_cases:
                self._write_test_case_row(ws, row, tc, req_text)
                row += 1

        # Настраиваем ширину колонок
        column_widths = [10, 40, 12, 30, 50, 40, 15, 25, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Автофильтр
        ws.auto_filter.ref = ws.dimensions

    def _write_test_case_row(self, ws, row: int, tc: TestCase, req_text: str):
        """Записывает строку тест-кейса."""
        # Форматируем предусловия
        preconditions = "\n".join(f"- {p}" for p in tc.preconditions) if tc.preconditions else ""

        # Форматируем шаги
        steps = "\n".join(f"{s['step']}. {s['action']}" for s in tc.steps) if tc.steps else ""

        # Санитизируем все значения для защиты от formula injection
        values = [
            sanitize_for_excel(tc.id),
            sanitize_for_excel(tc.title),
            sanitize_for_excel(tc.priority),
            sanitize_for_excel(preconditions),
            sanitize_for_excel(steps),
            sanitize_for_excel(tc.expected_result),
            sanitize_for_excel(tc.test_type),
            sanitize_for_excel(tc.technique),
            sanitize_for_excel(req_text)
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = self.CELL_ALIGNMENT
            cell.border = self.THIN_BORDER

            # Подсветка приоритета
            if col == 3 and value in self.PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=self.PRIORITY_COLORS[value],
                    end_color=self.PRIORITY_COLORS[value],
                    fill_type="solid"
                )

    def _create_analysis_sheet(self, ws, results: list[GenerationResult]):
        """Создает лист с анализом требований."""
        headers = ["Требование", "Входные данные", "Выходные данные", "Бизнес-правила", "Состояния"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        row = 2
        for result in results:
            analysis = result.analysis
            values = [
                result.requirement_text[:200] + "..." if len(result.requirement_text) > 200 else result.requirement_text,
                "\n".join(f"- {i}" for i in analysis.inputs),
                "\n".join(f"- {o}" for o in analysis.outputs),
                "\n".join(f"- {r}" for r in analysis.business_rules),
                "\n".join(f"- {s}" for s in analysis.states) if analysis.states else "N/A"
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.THIN_BORDER

            row += 1

        # Ширина колонок
        column_widths = [60, 40, 40, 50, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_summary_sheet(self, ws, results: list[GenerationResult], group_by_layer: bool = False):
        """Создает сводный лист."""
        # Заголовок
        ws.cell(row=1, column=1, value="Сводка по генерации тест-кейсов").font = Font(bold=True, size=14)

        # Статистика
        total_tests = sum(len(r.test_cases) for r in results)
        total_requirements = len(results)
        total_tokens = sum(r.tokens_used for r in results)

        stats = [
            ("Всего требований:", total_requirements),
            ("Всего тест-кейсов:", total_tests),
            ("Токенов использовано:", total_tokens),
        ]

        row = 3
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Распределение по слоям (если group_by_layer или есть данные о слоях)
        layer_counts = {"api": 0, "ui": 0, "integration": 0, "e2e": 0}
        component_counts = {"backend": 0, "frontend": 0, "fullstack": 0}

        for result in results:
            for tc in result.test_cases:
                layer = getattr(tc, 'layer', 'api') or 'api'
                component = getattr(tc, 'component', 'fullstack') or 'fullstack'
                if layer in layer_counts:
                    layer_counts[layer] += 1
                if component in component_counts:
                    component_counts[component] += 1

        # Показываем статистику по слоям если есть разнообразие или включена группировка
        if group_by_layer or sum(1 for c in layer_counts.values() if c > 0) > 1:
            row += 1
            ws.cell(row=row, column=1, value="Распределение по слоям:").font = Font(bold=True)
            row += 1

            for layer, count in layer_counts.items():
                if count > 0:
                    cell = ws.cell(row=row, column=1, value=layer.upper())
                    ws.cell(row=row, column=2, value=count)
                    row += 1

        # Распределение по компонентам
        if sum(1 for c in component_counts.values() if c > 0) > 1:
            row += 1
            ws.cell(row=row, column=1, value="Распределение по компонентам:").font = Font(bold=True)
            row += 1

            for component, count in component_counts.items():
                if count > 0:
                    cell = ws.cell(row=row, column=1, value=component.capitalize())
                    if component in self.COMPONENT_COLORS:
                        cell.fill = PatternFill(
                            start_color=self.COMPONENT_COLORS[component],
                            end_color=self.COMPONENT_COLORS[component],
                            fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF")
                    ws.cell(row=row, column=2, value=count)
                    row += 1

        # Распределение по приоритетам
        row += 1
        ws.cell(row=row, column=1, value="Распределение по приоритетам:").font = Font(bold=True)
        row += 1

        priority_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        for result in results:
            for tc in result.test_cases:
                if tc.priority in priority_counts:
                    priority_counts[tc.priority] += 1

        for priority, count in priority_counts.items():
            cell = ws.cell(row=row, column=1, value=priority)
            if priority in self.PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=self.PRIORITY_COLORS[priority],
                    end_color=self.PRIORITY_COLORS[priority],
                    fill_type="solid"
                )
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Распределение по техникам
        row += 1
        ws.cell(row=row, column=1, value="Распределение по техникам:").font = Font(bold=True)
        row += 1

        technique_counts = {}
        for result in results:
            for tc in result.test_cases:
                technique_counts[tc.technique] = technique_counts.get(tc.technique, 0) + 1

        for technique, count in sorted(technique_counts.items()):
            ws.cell(row=row, column=1, value=technique)
            ws.cell(row=row, column=2, value=count)
            row += 1

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15


class CSVExporter:
    """Экспортер в формат CSV."""

    def export(self, results: list[GenerationResult], output_path: str) -> str:
        """
        Экспортирует результаты в CSV файл.

        Args:
            results: Список результатов генерации
            output_path: Путь для сохранения

        Returns:
            Путь к созданному файлу
        """
        logger.info(f"Начало экспорта в CSV: {output_path}")
        path = Path(output_path)
        if not path.suffix:
            path = path.with_suffix(".csv")

        headers = [
            "ID", "Title", "Priority", "Preconditions",
            "Steps", "Expected Result", "Test Type", "Technique", "Requirement"
        ]

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
            writer.writerow(headers)

            for result in results:
                req_text = result.requirement_text[:100] + "..." if len(result.requirement_text) > 100 else result.requirement_text

                for tc in result.test_cases:
                    preconditions = " | ".join(tc.preconditions) if tc.preconditions else ""
                    steps = " | ".join(f"{s['step']}. {s['action']}" for s in tc.steps) if tc.steps else ""

                    # Санитизируем все значения для защиты от formula injection
                    row = [
                        sanitize_for_excel(tc.id),
                        sanitize_for_excel(tc.title),
                        sanitize_for_excel(tc.priority),
                        sanitize_for_excel(preconditions),
                        sanitize_for_excel(steps),
                        sanitize_for_excel(tc.expected_result),
                        sanitize_for_excel(tc.test_type),
                        sanitize_for_excel(tc.technique),
                        sanitize_for_excel(req_text)
                    ]
                    writer.writerow(row)

        logger.info(f"CSV файл сохранен: {path}")
        return str(path)


class AllureCSVExporter:
    """
    Экспортер в формат CSV для Allure TestOps.

    Формат соответствует реальному экспорту из Allure TestOps:
    - Разделитель: точка с запятой (;)
    - Сценарий: [step N] для шагов, \\t[expected N.1] для expected results
    - Поддержка кастомных полей через маппинг
    """

    # Стандартные заголовки Allure TestOps (без allure_id для новых тест-кейсов)
    HEADERS = [
        "name",             # Название (обязательное)
        "full_name",        # Полное имя
        "automated",        # Автоматизирован (true/false)
        "description",      # Описание
        "precondition",     # Предусловия
        "expected_result",  # Ожидаемый результат (на уровне тест-кейса)
        "status",           # Статус (Draft, Ready, etc.)
        "scenario",         # Сценарий с шагами
        "tag",              # Теги
        "link",             # Ссылки
        "example",          # Примеры
        "parameter",        # Параметры
        "Jira",             # Ссылка на Jira
        "Lead",             # Лид
        "Owner",            # Владелец
        "Suite",            # Сьют
        "Story",            # Story
        "Feature",          # Feature
        "Epic",             # Epic
    ]

    def export(
        self,
        results: list[GenerationResult],
        output_path: str,
        status: str = "Draft",
        suite: str = "",
        feature: str = "",
        epic: str = "",
        owner: str = "",
        jira_link: str = ""
    ) -> str:
        """
        Экспортирует результаты в CSV файл для Allure TestOps.

        Args:
            results: Список результатов генерации
            output_path: Путь для сохранения
            status: Статус тест-кейсов (Draft, Ready, etc.)
            suite: Название Suite для всех тест-кейсов
            feature: Название Feature
            epic: Название Epic
            owner: Владелец тест-кейсов
            jira_link: Ссылка на задачу в Jira

        Returns:
            Путь к созданному файлу
        """
        logger.info(f"Начало экспорта в Allure CSV: {output_path}")
        path = Path(output_path)
        if not path.suffix:
            path = path.with_suffix(".csv")

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(self.HEADERS)

            for result in results:
                for tc in result.test_cases:
                    row = self._format_test_case(
                        tc, result, status, suite, feature, epic, owner, jira_link
                    )
                    writer.writerow(row)

        logger.info(f"Allure CSV файл сохранен: {path}")
        return str(path)

    def _format_test_case(
        self,
        tc: TestCase,
        result: GenerationResult,
        status: str,
        suite: str,
        feature: str,
        epic: str,
        owner: str,
        jira_link: str
    ) -> list[str]:
        """Форматирует тест-кейс в формат Allure TestOps."""

        # Формируем теги
        tags = self._format_tags(tc)

        # Формируем сценарий в формате Allure
        scenario = self._format_scenario(tc.steps, tc.expected_result)

        # Предусловия - многострочный текст
        precondition = ""
        if tc.preconditions:
            # Нумеруем предусловия
            precondition_lines = []
            for i, p in enumerate(tc.preconditions, 1):
                precondition_lines.append(f"{i}. {p}")
            precondition = "\n".join(precondition_lines)

        # Описание - связь с требованием
        description = ""
        if result.requirement_text:
            req_text = result.requirement_text[:1000]
            description = f"Требование:\n{req_text}"

        # Формируем Story из техники тестирования
        story = tc.technique or ""

        row = [
            sanitize_for_excel(tc.title),          # name
            "",                                    # full_name
            "false",                               # automated
            sanitize_for_excel(description),       # description
            sanitize_for_excel(precondition),      # precondition
            sanitize_for_excel(tc.expected_result or ""),  # expected_result
            status,                                # status
            scenario,                              # scenario
            tags,                                  # tag
            "",                                    # link
            "",                                    # example
            "",                                    # parameter
            jira_link,                             # Jira
            "",                                    # Lead
            owner,                                 # Owner
            suite,                                 # Suite
            story,                                 # Story
            feature,                               # Feature
            epic,                                  # Epic
        ]

        return row

    def _format_scenario(self, steps: list[dict], expected_result: str = "") -> str:
        """
        Форматирует шаги в формат сценария Allure TestOps.

        Формат Allure:
        [step N] Действие
        \\t[expected N.1] Expected Result
        \\t\\t[expected.step N.1.1] Конкретный результат

        Args:
            steps: Список шагов [{step: int, action: str}, ...]
            expected_result: Общий ожидаемый результат

        Returns:
            Отформатированный сценарий
        """
        if not steps:
            return ""

        formatted_lines = []

        for step in steps:
            step_num = step.get('step', '')
            action = step.get('action', '')

            # Основной шаг
            formatted_lines.append(f"[step {step_num}] {action}")

            # Если есть expected для конкретного шага
            step_expected = step.get('expected', '')
            if step_expected:
                formatted_lines.append(f"\t[expected {step_num}.1] Expected Result")
                formatted_lines.append(f"\t\t[expected.step {step_num}.1.1] {step_expected}")

        # Добавляем общий expected result после последнего шага если есть
        if expected_result and steps:
            last_step = steps[-1].get('step', len(steps))
            # Проверяем, не добавили ли уже expected для последнего шага
            if not steps[-1].get('expected', ''):
                formatted_lines.append(f"\t[expected {last_step}.1] Expected Result")
                formatted_lines.append(f"\t\t[expected.step {last_step}.1.1] {expected_result}")

        return "\n".join(formatted_lines)

    def _format_tags(self, tc: TestCase) -> str:
        """
        Форматирует теги для Allure TestOps.

        Args:
            tc: Тест-кейс

        Returns:
            Строка тегов через запятую
        """
        tags = []

        # Добавляем существующие теги
        existing_tags = getattr(tc, 'tags', []) or []
        tags.extend(existing_tags)

        # Добавляем тип теста как тег
        if tc.test_type:
            tags.append(tc.test_type)

        # Добавляем приоритет как тег
        if tc.priority:
            tags.append(tc.priority)

        # Добавляем слой как тег
        layer = getattr(tc, 'layer', '')
        if layer:
            tags.append(layer.upper())

        return ",".join(tags)
